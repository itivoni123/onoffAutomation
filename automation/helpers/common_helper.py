import logging
import random
import socket
import itertools
from inspect import signature
from distutils.util import strtobool
from collections.abc import Iterable

from os import path, listdir, environ, makedirs, remove
from re import compile as recompile, sub
from csv import DictReader, reader
from time import sleep
from types import FunctionType
from string import hexdigits
from logging import info
from fnmatch import fnmatch
from hashlib import sha1 as hashlibsha1
from operator import ne
from datetime import datetime, date
from pydash import snake_case
from openpyxl import load_workbook
from pydash.strings import kebab_case
from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException

# from enums import DsarTopics, ExportReportTypes, StorageRepo
# from configurations.shared_consts import TIME_PRINTS_INTERVAL, MINUTE, AUTOMATION_WORKSPACE, TENANT_PROD_ID, \
#     TENANT_STAGING_ID, \
#     CVO_NAME_TO_WORKSPACE_ID
# from configurations.provider_related import Provider, AUTOMATION_ACCOUNT_NAME

# region Members
TIME_PRINTS_INTERVAL = 7
# DATA_SUBJECT_NAME_IN_DSAR_REPORT_IDX = 1
# ORGANIZATION_NAME_IN_DSAR_REPORT_IDX = 3
# DEAR_DATA_SUBJECT_IN_DSAR_REPORT_IDX = 5
# IDENTIFIED_FILES_IN_DSAR_REPORT_IDX = 7
# KEYS_TO_MASK = ['Password', 'password', 'Authorization', 'client_id', '--proxy-user', 'PAT']


# endregion Members

def move_hover_over_element(browser, by, element):
    ActionChains(browser).move_to_element(browser.find_element(by, element)).perform()


def switch_to_window_browser(browser, close_current=True):
    """
    Switch the focus to the next window
    :param browser: open the browser by the webdriver
    :param close_current:
    :return:
    """

    other_handle = None
    try:
        current_hundle = browser.current_window_handle
        other_handles = [
            h for h in browser.window_handles
            if h != current_hundle
        ]
        sleep(2)
        if len(other_handles) == 0:
            other_handle = current_hundle
        else:
            other_handle = other_handles[0]

        # switch to the other window to close it
        if close_current:
            browser.close()
        browser.switch_to_window(other_handle)
    except NoSuchElementException:
        if other_handle >= 1:
            switch_to_window_browser(browser, close_current=False)
        else:
            return True
    return True


def execute_js(browser, js_code):
    return browser.execute_script(js_code)


def is_download_done(downloaded_file_path, wait_time_secs=30, expected_hash="", block_size=2 ** 20):
    starting_to_wait = datetime.now()
    while (datetime.now() - starting_to_wait).seconds < wait_time_secs:
        try:
            assert path.isfile(downloaded_file_path)
            return True
        except AssertionError:
            pass
    if expected_hash:
        validate_hash(downloaded_file_path, expected_hash, block_size)
    return False


def validate_hash(file_to_check, hash_to_check, block_size=2 ** 20):
    with open(file_to_check, 'rb') as frb:
        sha1 = hashlibsha1()
        buf = frb.read(block_size)
        while len(buf) > 0:
            sha1.update(buf)
            buf = frb.read(block_size)
        assert sha1.hexdigest() == hash_to_check, (sha1.hexdigest(), hash_to_check)


def generate_random_name_with_prefix(prefix_text=''):
    random_seed = int(environ.get('BUILD_NUMBER', '1'))
    random_seed += datetime.now().microsecond
    random.seed(random_seed)

    return prefix_text + ' '.join([random.choice(hexdigits) for _ in range(8)])


def get_int_from_any_string(str_to_int):
    """
    Using regex to remove any character that is not a number and convert to int
    WILL RETURN THE ALL NUMBER HE FINDS IN A LIST - "GF12GFB56" will return [12, 56]
    :param str_to_int: any String (w/o numbers)
    :return: Only the numbers in a list
    """
    numbers_found = list()
    numbers_regex = recompile(r'\d+')
    for regex_match in numbers_regex.finditer(str_to_int):
        numbers_found.append(int(regex_match.group()))
    return numbers_found


def get_first_int_from_string(main_string, starts_with='', ends_with=''):
    if not starts_with or not ends_with:
        return get_int_from_any_string(main_string)[0]
    return get_int_from_any_string(find_string_between(main_string, starts_with, ends_with))[0]


def find_string_between(main_string, starts_with, ends_with):
    try:
        start_index = main_string.index(starts_with) + len(starts_with)
        end_index = main_string.index(ends_with, start_index)
        return main_string[start_index:end_index]
    except ValueError:
        return ""


def get_text_of_docx_file(doc_file):
    text_found = []

    docx_file = Document(doc_file)

    for paragraph in docx_file.paragraphs:
        paragraph_text = paragraph.text
        if not paragraph_text:
            paragraph_text = get_hyperlink_from_docx_paragraph(paragraph)
        if paragraph_text and not paragraph_text == '\n':
            text_found.append(paragraph_text.strip())

    return text_found


def get_hyperlink_from_docx_paragraph(paragraph):
    xml = paragraph._element.xml
    xml = find_string_between(xml, "<w:hyperlink ", "</w:hyperlink>")
    return find_string_between(xml, "<w:t>", "</w:t>")


def get_dsar_topic_value(dsar_value):
    dsar_value = dsar_value if dsar_value else ""
    dsar_value = dsar_value[0] if isinstance(dsar_value, list) and len(dsar_value) == 1 else dsar_value
    return dsar_value


def get_data_subject_name_and_organization(data_subject_text, organization_text):
    data_subject_key, data_subject_name = data_subject_text.split(":")
    organization_key, organization_name = organization_text.split(":")
    return {
        data_subject_key: data_subject_name.strip(),
        organization_key: organization_name.strip()
    }


def shorten_number_to_full_number(number_with_string):
    """
        This function should be use an any place that the number can be higher than 1,000

    :param number_with_string: 1.5k
    :return:
    """
    just_the_numbers = get_int_from_any_string(number_with_string)
    full_number = just_the_numbers[0]

    if '.' in number_with_string:
        full_number += float((just_the_numbers[1] / 10))
    if 'k' in number_with_string.lower():
        full_number = full_number * 1000
    if 'm' in number_with_string.lower():
        full_number = full_number * 1000 * 1000
    return int(full_number)


def convert_variable_to(variable, variable_type=None):
    variable_type = variable_type or list
    if isinstance(variable, variable_type):
        return variable
    if isinstance(variable, Iterable) and not isinstance(variable, str) and not isinstance(variable, dict):
        return variable_type(variable)

    if variable_type == list:
        return [variable]
    if variable_type == set:
        return {variable}
    if variable_type == tuple:
        return tuple(variable)

    assert "Missing Conversion Type"


def run_method_in_while(max_wait_time, method, *args, **kwargs):
    return run_method_until(max_wait_time, method, method_done_if_ne_to=False, *args, **kwargs)


def run_method_until_not_none(max_wait_time, method, *args, **kwargs):
    return run_method_until(max_wait_time, method, method_done_if_ne_to=None, *args, **kwargs)


def run_method_until(max_wait_time, method, *args, **kwargs):
    try:
        method_params = list(signature(method).parameters.keys())
    except (ValueError, TypeError):
        method_params = []

    if 'print_logs' not in method_params:
        should_print_logs = kwargs.pop('print_logs', True)
    else:
        should_print_logs = kwargs.get('print_logs', True)
    api_delay = int(kwargs.pop('api_delay', '0'))
    method_done_if_ne_to = kwargs.pop('method_done_if_ne_to', False)
    text_to_print = 'The method: "%s" is still RUNNING(%s/%s)!! args:(%s) ****** kwargs:(%s)\nCurrent method result: %s'
    log_printed = True
    start_time = datetime.now()
    method_running_time = (datetime.now() - start_time).seconds

    while method_running_time < max_wait_time:
        is_method_done = method(*args, **kwargs)
        if should_print_logs:
            log_printed = print_every_x_seconds(method_running_time, log_printed, text_to_print % (
                method.__name__, method_running_time, max_wait_time, args, kwargs, is_method_done))
        if ne(is_method_done, method_done_if_ne_to):
            if should_print_logs:
                info('The method: "%s" FINISHED running(%s/%s)!! args:(%s) ##### kwargs:(%s)' % (
                    method.__name__, method_running_time, max_wait_time, args, kwargs))
            return is_method_done
        method_running_time = (datetime.now() - start_time).seconds
        sleep(api_delay)
    assert False, 'The method: "%s" has TIMED OUT(%s)!!' % (method.__name__, max_wait_time)


def print_every_x_seconds(method_running_time, should_print, text_to_print):
    if method_running_time % TIME_PRINTS_INTERVAL == 0:
        if not should_print:
            info(text_to_print)
        return True
    else:
        return False


def get_data_from_csv(file_path):

    data_list = list()
    with open(f'albums/{file_path}') as csv_file:
        for row in DictReader(csv_file):
            data_list.append(row)
    return data_list


def get_column_from_csv(file_path, column_name):
    column_data = []

    info("Going through '%s', trying to extract the column: '%s' " % (file_path, column_name))

    with open(file_path) as csv_file:
        for row in DictReader(csv_file):
            try:
                column_data.append(row[column_name])
            except KeyError:
                assert False, "Key doesn't exists, Keys found: %s" % row.keys()

    info("Finished column extraction from '%s', column: '%s' " % (file_path, column_name))
    return column_data


def get_row_by_file_name_from_csv(file_path, file_name_to_search, with_last_modified=False, with_duplicates=False,
                                  with_created=False, with_last_accessed=False):
    info("Going through '%s', trying to extract the row for the file name: '%s' " % (file_path, file_name_to_search))

    with open(file_path) as csv_file:
        for row in DictReader(csv_file):
            if row['File Name'] == file_name_to_search:
                files_row = dict(row)
                files_row['Personal Information'] = get_patterns_from_string(files_row['Personal Information'])
                files_row['Sensitive Personal Information'] = \
                    get_patterns_from_string(files_row['Sensitive Personal Information'])
                info("Finished column extraction from '%s', column: '%s' " % (file_path, file_name_to_search))

                if not with_last_modified:
                    files_row.pop('Last Modified', '')
                if not with_created:
                    files_row.pop('Created Time', '')
                if not with_last_accessed:
                    files_row.pop('Last Accessed', '')
                if not with_duplicates:
                    files_row.pop('Duplicates', '')

                return files_row


def get_patterns_from_string(patterns_string):
    if not patterns_string:
        return ""

    patterns_dict = {}
    for single_pattern_string in patterns_string.split(','):
        pattern = single_pattern_string.split(":")
        patterns_dict[pattern[0]] = pattern[1]

    return patterns_dict


def convert_to_kebab(text, is_occm_kebab=False):
    if is_occm_kebab:
        text = sub('([A-Z][a-z]+)(!([A-Za-z0-9]))', r'\1-\2', text)
        text = sub('([a-z0-9])([A-Z])', r'\1-\2', text)
        text = sub(r'(\d+)([A-Za-z])', r'\1-\2', text)
        text = sub('([a-zA-Z])([0-9]{2,})', r'\1-\2', text)
        return text.lower().replace(' - ', '-').replace('_', '-').replace(' ', '-')

    else:
        return kebab_case(text)


def to_camel_case(snake_str):
    components = snake_str.split('-')
    # We capitalize the first letter of each component except the first one
    # with the 'title' method and join them together.
    return components[0] + ''.join(x.title() for x in components[1:])


def get_running_build():
    agent = environ.get('AGENT', '')
    build_num = environ.get('BUILD_NUMBER', '0')
    if 1 < int(build_num):
        agent = build_num
    elif not agent:
        agent = socket.gethostname()

    assert agent, "AGENT env could not be set"
    return agent


def get_running_agent():
    agent = environ.get('AGENT', '')
    if not agent:
        agent = socket.gethostname()
    assert agent, "AGENT env could not be set"
    return agent


def get_creds_file(env_type, cred_type='cvo'):
    file_content_list = []
    creds_path = f"/{env_type}/{cred_type}"
    if path.exists("/var" + creds_path):
        creds_path = "/var" + creds_path

    if path.exists(creds_path):
        with open(creds_path, "r") as f:
            line = f.readline().strip()
            while line:
                file_content_list.append(line)
                line = f.readline().strip()
    else:
        logging.error("[ X ] Path doesn't exists: " + creds_path)
    return file_content_list


def get_user_cred(user_file_name):
    username, password = get_creds_file(user_file_name)
    return username, password


def add_cvo_for_deletion(env, cvo_name):
    cvos_path = "/test_cvos"
    if not path.exists(cvos_path):
        cvos_path = "/tmp/%s" % cvos_path

    if not path.exists(cvos_path):
        makedirs(cvos_path)

    with open("%s/%s_%s" % (cvos_path, env, cvo_name), 'w') as f:
        f.write("temp file")


def remove_cvo_from_files_list(env, cvo_name):
    cvos_path = "/test_cvos/%s_%s" % (env, cvo_name)
    if not path.exists(cvos_path):
        cvos_path = "/tmp" + cvos_path

    if path.exists(cvos_path):
        remove(cvos_path)


def get_cvo_from_local_files_list():
    cvos_path = "/test_cvos/"
    if not path.exists(cvos_path):
        cvos_path = "/tmp" + cvos_path
    if not path.exists(cvos_path):
        return []

    return [cvos_path + file_path for file_path in listdir(cvos_path)]


def get_occm_token_client_id():
    return get_creds_file(get_running_environment(), "token_client_id")[0]


def get_multi_ids_to_css_selector(*id_selectors):
    selectors = []
    css_selector = ""

    if not id_selectors:
        return ""

    for selector in id_selectors:
        if not selector:
            continue
        if isinstance(selector, str):
            selectors.append(selector)
        if isinstance(selector, list):
            selectors += selector

    for selector in selectors:
        css_selector += f"[id={selector}]"
        if selector != selectors[-1]:
            css_selector += ","

    return css_selector


def get_running_environment():
    return "staging" if var_to_bool(environ.get('IS_STAGING', True)) else "prod"


def var_to_bool(variable):
    if type(variable) == bool:
        return variable
    return bool(strtobool(f"{variable}".replace("'", "").replace('"', "")))


def get_mount_filepath_by_storage_type(storage_type, volume_name, file_path=''):
    info(f"{storage_type}, {volume_name}, {file_path}")
    mount_path = f"{file_path.lstrip('/')}"
    volume_name = volume_name.lstrip('/')
    storage_type = storage_type.upper()
    if storage_type == "ANF":
        mount_path = path.join(volume_name.replace('_', '-'), mount_path)
    elif storage_type != "S3":
        mount_path = path.join(volume_name, mount_path)
    return path.realpath(f"/{mount_path}")


def list_methods(cls):
    return set(x for x, y in cls.__dict__.items()
               if isinstance(y, (FunctionType, classmethod, staticmethod)))


def list_parent_methods(cls):
    return set(itertools.chain.from_iterable(
        list_methods(c).union(list_parent_methods(c)) for c in cls.__bases__))


def list_subclass_methods(cls, is_narrow):
    methods = list_methods(cls)
    if is_narrow:
        parent_methods = list_parent_methods(cls)
        return set(cls for cls in methods if not (cls in parent_methods))
    else:
        return methods


def get_random_storage():
    return environ.get("RANDOM_STORAGE")


def get_xslx_without_headers_in_dict(file_path):
    data = {}
    ws = load_workbook(file_path)
    for row in ws.worksheets[0].iter_rows():
        file_name = f"{row[0].value}".strip()
        data[file_name] = sorted([f"{row_data.value}".strip() for row_data in row[1:] if row_data.value is not None])
    return data


def str_last_replace(source_string, what_to_replace, replace_with):
    start_str, removed_str, end_str = source_string.rpartition(what_to_replace)
    return start_str + replace_with + end_str


def word_to_single_case(word):
    if word.endswith("ies"):
        return f"{word[:-3]}y"
    if word.endswith("s"):
        return word[:-1]
    return word


def assert_item_approx(expected_value, found_value, proximity_range) -> bool:
    if expected_value == found_value:
        return True
    min_range = int(expected_value) - proximity_range
    max_range = int(expected_value) + proximity_range
    return found_value in range(min_range, max_range)


def merge_dicts(dict1, dict2):
    tmp_dict = dict(**dict1)
    for key, value in dict2.items():
        value_in_dict1 = dict1.get(key)
        is_value_bool = type(dict1.get(key)) == bool
        tmp_dict[key] = value + value_in_dict1 if value_in_dict1 and not is_value_bool else value
    return tmp_dict


def normalize_string(text):
    return text.replace('-', '–')


def get_current_date():

    # Textual month, day and year
    today = date.today()
    d2 = today.strftime("%B %d, %Y").split()

    # Remove the year
    d2.pop()

    # Combines between the month and day
    crr_date = d2[0] + '_' + d2[1].replace(",", "")
    return crr_date
